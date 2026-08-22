"""
Generate the joint-issuance authority classification audit.

Classification basis
====================
We classify each named issuing authority into one of EIGHT mutually
exclusive types, jointly defined by two dimensions that are explicit
in Chinese constitutional and party-law texts:

  Dim-1  Administrative level
    - Central (中央):    issued from Beijing, jurisdiction = nation
    - Provincial (省级): 省 / 直辖市 / 自治区 / 特别行政区
    - Municipal (地市级): 设区的市 / 自治州 / 盟

  Dim-2  Functional system
    - Party (党的系统):       Communist Party committees, party
                              working departments, party-led mass
                              organisations (共青团, 工会, 妇联).
                              [Reference: 中国共产党章程, Chs III-VI]
    - Government (政府系统):   人民政府 + 政府工作部门 (厅/局/委).
                              [Reference: 宪法 §85, §105 - §107]
    - Science / Society (学会): 中科院 / 工程院 / 社科院 / 中国科协 /
                              省级科协 — sui generis academic and
                              learned bodies, treated as a distinct
                              functional system in policy analysis.

The 2 × 3 grid plus three special national bodies yields the following
8 categories used in panel (b) of Figure 4.1:

  1. Central Party System         (Cat-1)
  2. State Council & Ministries   (Cat-2)
  3. National Sci. / Society      (Cat-3)
  4. Provincial Party             (Cat-4)
  5. Provincial Gov. & Bureau     (Cat-5)
  6. Municipal Party              (Cat-6)
  7. Municipal Gov. & Bureau      (Cat-7)
  8. Other                        (Cat-8)

This 8-fold scheme is parallel in granularity to the 8 open-science
categories used in panel (a), enabling an apples-to-apples reading of
the two co-occurrence matrices.
"""
from pathlib import Path
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

ROOT       = Path(__file__).resolve().parent.parent
FILES_XLSX = ROOT / "02_data" / "final_entries" / "政策文件表_v5.3.xlsx"

# ---- 8-category classification ----
TYPE_ORDER = [
    "Central Party",
    "State Council & Ministries",
    "National Sci./Society",
    "Provincial Party",
    "Provincial Gov./Bureau",
    "Municipal Party",
    "Municipal Gov./Bureau",
    "Other",
]
TYPE_ZH = {
    "Central Party":              "中央党的系统",
    "State Council & Ministries": "国务院与中央部委",
    "National Sci./Society":      "国家科研学会",
    "Provincial Party":           "省级党委系统",
    "Provincial Gov./Bureau":     "省级政府与厅局",
    "Municipal Party":            "地市党委系统",
    "Municipal Gov./Bureau":      "地市政府与厅局",
    "Other":                      "其他",
}

# National-level science / society / mass-academic bodies
NATIONAL_SCI = {
    "中国科学院", "中国工程院", "中国社会科学院",
    "中国科学技术协会", "中国科协",
    "国家自然科学基金委员会", "国家自然科学基金",
    "中国气象局",  # 副部级事业单位
}
# Central party-led mass organisations
PARTY_MASS_NATIONAL = {
    "共青团中央", "中华全国总工会", "全国妇女联合会",
    "中央军事委员会", "中央军委",
}

def _strip(n: str) -> str:
    # Drop parenthesised suffixes such as (已撤销) / (原XX) etc.
    return re.sub(r"[(（][^)）]*[)）]", "", str(n)).strip()

def _has_local_marker(n: str):
    """Return ('central'|'provincial'|'municipal') based on prefix."""
    prov_pat = r"(省|自治区|北京市|天津市|上海市|重庆市|特别行政区)"
    pref_pat = r"(地区|州|盟)"
    if re.search(prov_pat, n):
        return "provincial"
    # A 'X市' that is NOT one of the 4 municipalities = prefecture-level city
    if re.search(r"市", n) and not re.search(r"(北京市|天津市|上海市|重庆市)", n):
        return "municipal"
    if re.search(pref_pat, n):
        return "municipal"
    return "central"

def classify(name) -> str:
    if not isinstance(name, str) or not name.strip():
        return "Other"
    n = _strip(name)

    # National sci/society/academy
    if any(k in n for k in NATIONAL_SCI):
        return "National Sci./Society"
    if n in PARTY_MASS_NATIONAL:
        return "Central Party"

    is_party = (n.startswith("中共") or "中国共产党" in n or
                "党委" in n or "党组" in n or n.startswith("中央"))

    level = _has_local_marker(n)

    if level == "central":
        if is_party:
            return "Central Party"
        if "国务院" in n:
            return "State Council & Ministries"
        # National admin bodies
        if (n.endswith("部") or n.endswith("总局") or n.endswith("总署") or
            n.endswith("银行") or n.endswith("管理局") or
            n.endswith("委员会") or n.startswith("国家") or
            n.startswith("中国") or n.startswith("全国")):
            return "State Council & Ministries"
        return "Other"

    if level == "provincial":
        if is_party:
            return "Provincial Party"
        return "Provincial Gov./Bureau"

    if level == "municipal":
        if is_party:
            return "Municipal Party"
        return "Municipal Gov./Bureau"

    return "Other"

# ---- Build audit table ----
files = pd.read_excel(FILES_XLSX)
joint = files[files["is_joint_issuance"] == True].copy().reset_index(drop=True)

def split_auth(s):
    if pd.isna(s): return []
    return [x.strip() for x in str(s).split(";") if x.strip()]

joint["authorities"] = joint["issuing_authority"].apply(split_auth)
joint["n_authorities"] = joint["authorities"].str.len()
joint["authority_types"] = joint["authorities"].apply(
    lambda L: [classify(a) for a in L])
joint["type_set"] = joint["authority_types"].apply(
    lambda L: sorted({t for t in L if t != "Other"}))
joint["type_set_str"] = joint["type_set"].apply(lambda L: " | ".join(L))
joint["authority_pretty"] = joint["authorities"].apply(lambda L: " | ".join(L))
joint["types_pretty"]    = joint["authority_types"].apply(lambda L: " | ".join(L))

audit = joint[[
    "policy_file_id", "issue_year", "governance_level", "central_local",
    "policy_title", "n_authorities",
    "authority_pretty", "types_pretty", "type_set_str",
]].rename(columns={
    "policy_file_id":   "file_id",
    "issue_year":       "year",
    "governance_level": "gov_level",
    "policy_title":     "title",
    "authority_pretty": "authorities (raw)",
    "types_pretty":     "authority types (1-1 mapping)",
    "type_set_str":     "type set (unique)",
})

# Distribution table
from collections import Counter
typ_counter = Counter()
for L in joint["authority_types"]:
    typ_counter.update(L)
dist = (pd.DataFrame(
    [(TYPE_ZH[t], t, typ_counter[t]) for t in TYPE_ORDER],
    columns=["类型 (中文)", "Type (English)", "Authority mentions"]
        )
        .assign(percent=lambda d: d["Authority mentions"]
                / d["Authority mentions"].sum() * 100))

# Co-occurrence and Jaccard
import numpy as np
n_t = len(TYPE_ORDER)
B = np.zeros((len(joint), n_t), dtype=bool)
for r, ts in enumerate(joint["type_set"]):
    for t in ts:
        if t in TYPE_ORDER:
            B[r, TYPE_ORDER.index(t)] = True

co = np.zeros((n_t, n_t), dtype=int)
jac = np.zeros((n_t, n_t), dtype=float)
for i in range(n_t):
    for j in range(n_t):
        inter = int((B[:, i] & B[:, j]).sum())
        union = int((B[:, i] | B[:, j]).sum())
        co[i, j]  = inter
        jac[i, j] = inter / union if union > 0 else 0.0

co_df  = pd.DataFrame(co,  index=TYPE_ORDER, columns=TYPE_ORDER)
jac_df = pd.DataFrame(jac.round(4),
                      index=TYPE_ORDER, columns=TYPE_ORDER)

# ---- Append new sheets to 政策文件表_v5.3.xlsx ----
new_sheets = {
    "联合发文-分类说明":    None,  # populated below
    "联合发文-文件清单":    audit,
    "联合发文-类型分布":    dist,
    "联合发文-共现矩阵":    co_df.reset_index().rename(columns={"index": "type"}),
    "联合发文-Jaccard矩阵": jac_df.reset_index().rename(columns={"index": "type"}),
}

# Build the rationale sheet manually
rationale_rows = [
    ["联合发文部门分类标准 (v5.3 审计)"],
    [""],
    ["1. 分类依据"],
    ["   (1) 行政级别 — 依据《中华人民共和国宪法》第85-111条划分"],
    ["       中央 / 省级 / 地市级"],
    ["   (2) 功能系统 — 依据《中国共产党章程》第三-六章及党政机构改革文件"],
    ["       党的系统 / 政府系统 / 学会系统"],
    [""],
    ["2. 八个分类"],
]
for t in TYPE_ORDER:
    rationale_rows.append([f"   {t}", TYPE_ZH[t]])
rationale_rows += [
    [""],
    ["3. 规则要点"],
    ["   - 名称中先去除括号（如 (已撤销)、(原XX)），再判断"],
    ["   - 国家级学会与中科院/工程院/社科院归 National Sci./Society"],
    ["   - 共青团中央 / 全国总工会 / 全国妇联 归 Central Party"],
    ["   - 省级和地市级以名称前缀判断（直辖市归省级）"],
    [f"   - 联合发文文件总数: {len(joint)} 份"],
    [f"   - 涉及部门称次: {sum(joint['n_authorities'])} 次"],
    [f"   - 无法归类（Other）: {typ_counter['Other']} 次 "
     f"({typ_counter['Other']/sum(joint['n_authorities'])*100:.2f}%)"],
]
rationale_df = pd.DataFrame(rationale_rows)

# Write to the existing workbook (preserve original sheet)
wb = load_workbook(FILES_XLSX)
# Remove any old audit sheets first
for sn in list(new_sheets):
    if sn in wb.sheetnames:
        del wb[sn]

# Add rationale sheet
ws = wb.create_sheet("联合发文-分类说明")
for row in dataframe_to_rows(rationale_df, index=False, header=False):
    ws.append(row)

# Add other sheets
for sn, df in new_sheets.items():
    if df is None:
        continue
    ws = wb.create_sheet(sn)
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

wb.save(FILES_XLSX)
print(f"Appended {len(new_sheets)} audit sheets to {FILES_XLSX.name}")
print(f"  joint files:        {len(joint):,}")
print(f"  authority mentions: {sum(joint['n_authorities']):,}")
print("  type distribution:")
for t in TYPE_ORDER:
    print(f"    {t:30s} {typ_counter[t]:5d}  "
          f"({typ_counter[t]/sum(joint['n_authorities'])*100:5.1f}%)")
